'BY-SA-CC'

""" Automated Check for the
    /In fine print Section for drop ship 2, it is showing 7 business days instead of 12 business days/
    Issue wwhich will be directly written in the separate spreadsheet

    Given Permalinks & the Dropship Counts

    Needs testing on the No.Of Rows of the Input Spreadsheet. .!
"""


from twill import *
from bs4 import BeautifulSoup
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl import load_workbook


wb = load_workbook(filename = 'c:/users/c_thv/desktop/formula.xlsx')
ws1 = wb.get_sheet_by_name(name = 'eod')
endCol   = ws1.get_highest_column() #index of last used column, from 1
endRow   = ws1.get_highest_row()


#Index shouldn't be out of range when initiated through the xrange.
for i in range(1,endRow+1): 
    m = 'A' + str(i)
    n = 'B' + str(i)
    t = 'C' + str(i)
    m1 = ws1.cell(m).value
    mybrowser = get_browser()
    #print 'checking' + m1
    mybrowser.go(m1)
    soup = BeautifulSoup(mybrowser.get_html())
    mydivs = soup.find("div", class_= "fine-print").text
    if '12 business days' in mydivs:
        thilip = ws1.cell(n).value
        if thilip == 1:
            ws1.cell(t).value = 'Raise error as in Fine Print'
    else:
        print 'Got it'

wb.save("c:/users/c_thv/desktop/thilip.xlsx")


#if __name__ == '__main__':
 #   main()
